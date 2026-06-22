"""Excel 报表生成 - 使用 openpyxl 创建 Excel 工作簿"""

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def create_excel_report(
    data: list[dict[str, Any]] | dict[str, Any],
    output_path: str,
    sheet_name: str = "数据",
) -> str:
    """将数据写入 Excel 文件。

    Args:
        data: 数据源，支持字典列表或单层字典
        output_path: 输出文件路径
        sheet_name: 工作表名称

    Returns:
        输出文件的绝对路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 统一为列表格式
    rows = data if isinstance(data, list) else [data]

    if not rows:
        wb.save(output_path)
        return str(Path(output_path).resolve())

    # 写入表头
    headers = list(rows[0].keys())
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="000000"),
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 写入数据行
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row_data.get(header, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 自动调整列宽
    for col_idx, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 4, 50)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    return str(output.resolve())
